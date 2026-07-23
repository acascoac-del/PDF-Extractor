"""Generador de Excel (.xlsx).

Para facturas: hoja "Detalle" completa + hoja "Ítems" + hoja "Resumen".
Para tablas: una hoja por tabla con filas/columnas reales.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.document import Document, DocType
from app.models.extraction import Extraction


def generate_excel(doc: Document, extraction: Extraction) -> io.BytesIO:
    buf = io.BytesIO()
    wb = Workbook()

    if doc.doc_type == DocType.INVOICE and extraction:
        _write_invoice(wb, extraction)
    elif doc.doc_type == DocType.TABLE and extraction:
        _write_tables(wb, extraction)
    elif doc.doc_type == DocType.CONTRACT and extraction:
        _write_contract(wb, extraction)
    else:
        _write_generic(wb, doc, extraction)

    wb.save(buf)
    buf.seek(0)
    return buf


# ── Estilos ──────────────────────────────────────────────────────────────

def _header_font() -> Font:
    return Font(bold=True, size=11, color="FFFFFF")


def _section_font() -> Font:
    return Font(bold=True, size=12, color="1D4ED8")


def _header_fill() -> PatternFill:
    return PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")


def _light_fill() -> PatternFill:
    return PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")


def _thin_border() -> Border:
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()


def _section_row(ws, row: int, title: str, cols: int = 2) -> int:
    """Escribe una fila de sección con fondo azul claro."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _section_font()
    cell.fill = _light_fill()
    return row + 1


def _field_row(ws, row: int, label: str, value, col_label: int = 1, col_value: int = 2) -> int:
    """Escribe un campo label→valor."""
    if value is None or value == "":
        return row
    ws.cell(row=row, column=col_label, value=label).font = Font(bold=True, size=10)
    ws.cell(row=row, column=col_label).border = _thin_border()
    c = ws.cell(row=row, column=col_value, value=value)
    c.border = _thin_border()
    c.alignment = Alignment(wrap_text=True)
    return row + 1


def _f(fields: dict, key: str):
    """Obtiene el valor de un campo extraído."""
    return fields.get(key, {}).get("value")


# ── Factura ──────────────────────────────────────────────────────────────

def _write_invoice(wb: Workbook, ext: Extraction) -> None:
    fields = ext.data.get("fields", {})
    items = ext.data.get("items", [])
    ibp = ext.data.get("ibp_entries", [])
    tasas = ext.data.get("tasa_vial_entries", [])
    riesgos = ext.data.get("risk_descriptions", [])

    # ═══════════════════════════════════════════════════════════════════
    #  HOJA 1: DETALLE COMPLETO
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Detalle"
    ws.sheet_properties.tabColor = "1D4ED8"

    r = 1
    # ── Emisor ──
    r = _section_row(ws, r, "EMISOR", 2)
    for key, label in [
        ("emitter_name", "Razón Social"), ("emitter_cuit", "CUIT"),
        ("emitter_address", "Domicilio"), ("emitter_iibb", "Ingresos Brutos"),
        ("emitter_iva_condition", "Condición IVA"), ("emitter_start_date", "Inicio Actividades"),
    ]:
        r = _field_row(ws, r, label, _f(fields, key))

    # ── Receptor ──
    r = _section_row(ws, r, "RECEPTOR", 2)
    for key, label in [
        ("receptor_name", "Razón Social"), ("receptor_cuit", "CUIT"),
        ("receptor_address", "Domicilio"), ("receptor_iibb", "Ingresos Brutos"),
        ("receptor_account", "N° Cuenta"), ("receptor_deudor_account", "Cuenta Deudora"),
    ]:
        r = _field_row(ws, r, label, _f(fields, key))

    # ── Datos de la Factura ──
    r = _section_row(ws, r, "DATOS DE LA FACTURA", 2)
    for key, label in [
        ("invoice_number", "N° Factura"), ("invoice_letter", "Tipo (A/B/C)"),
        ("emission_date", "Fecha Emisión"), ("cae", "CAE"),
        ("invoice_cae", "CAE (pie)"), ("cae_expiry", "Vencimiento CAE"),
        ("iva_condition", "Condición IVA Receptor"),
        ("incoterms", "Incoterms"), ("sap_number", "N° SAP"),
        ("oc_number", "Orden de Compra"), ("payment_terms", "Condiciones de Pago"),
        ("shipping_method", "Forma de Envío"),
    ]:
        r = _field_row(ws, r, label, _f(fields, key))

    # ── Totales ──
    r = _section_row(ws, r, "TOTALES", 2)
    for key, label in [
        ("importe_neto", "Importe Neto"), ("financiacion", "Financiación"),
        ("icl_amount", "ICL"), ("idc_amount", "IDC"),
        ("iva_inscripto", "IVA Inscripto"), ("iva_no_inscripto", "IVA No Inscripto"),
        ("iva_percepcion", "IVA Percepción"), ("iva_percentage", "% IVA"),
        ("ingresos_brutos", "Importe Ing. Brutos"), ("tasa_vial", "Total Tasa Vial"),
        ("net", "Neto Gravado"), ("iva_amount", "IVA"), ("total", "TOTAL"),
    ]:
        r = _field_row(ws, r, label, _f(fields, key))

    # ── IBP (Ingresos Brutos Provincial) ──
    if ibp:
        r = _section_row(ws, r, "INGRESOS BRUTOS PROVINCIALES (IBP)", 3)
        _set_header(ws, r, ["Jurisdicción", "Alícuota %", "Importe"])
        r += 1
        for entry in ibp:
            ws.cell(row=r, column=1, value=entry.get("jurisdiction", "")).border = _thin_border()
            ws.cell(row=r, column=2, value=entry.get("percentage")).border = _thin_border()
            ws.cell(row=r, column=3, value=entry.get("amount")).border = _thin_border()
            r += 1

    # ── Tasa Vial ──
    if tasas:
        r = _section_row(ws, r, "TASA VIAL POR LOCALIDAD", 2)
        _set_header(ws, r, ["Localidad", "Importe"])
        r += 1
        for entry in tasas:
            ws.cell(row=r, column=1, value=entry.get("locality", "")).border = _thin_border()
            ws.cell(row=r, column=2, value=entry.get("amount")).border = _thin_border()
            r += 1

    # ── Riesgo/ONU ──
    if riesgos:
        r = _section_row(ws, r, "RIESGO / ONU", 2)
        _set_header(ws, r, ["N° Riesgo", "Descripción"])
        r += 1
        for entry in riesgos:
            ws.cell(row=r, column=1, value=entry.get("number", "")).border = _thin_border()
            ws.cell(row=r, column=2, value=entry.get("description", "")).border = _thin_border()
            r += 1

    # Ajustar anchos
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    # ═══════════════════════════════════════════════════════════════════
    #  HOJA 2: ÍTEMS (tabla completa)
    # ═══════════════════════════════════════════════════════════════════
    ws_items = wb.create_sheet("Ítems")
    ws_items.sheet_properties.tabColor = "16A34A"

    item_headers = ["Código", "Descripción", "Cantidad", "UM", "Riesgo ONU",
                    "Valor Unitario", "Importe", "Modificadores"]
    _set_header(ws_items, 1, item_headers)

    for i, item in enumerate(items, 2):
        ws_items.cell(row=i, column=1, value=item.get("code", "")).border = _thin_border()
        ws_items.cell(row=i, column=2, value=item.get("description", "")).border = _thin_border()
        ws_items.cell(row=i, column=3, value=item.get("quantity")).border = _thin_border()
        ws_items.cell(row=i, column=4, value=item.get("unit", "")).border = _thin_border()
        ws_items.cell(row=i, column=5, value=item.get("risk_un")).border = _thin_border()
        ws_items.cell(row=i, column=6, value=item.get("unit_price")).border = _thin_border()
        ws_items.cell(row=i, column=7, value=item.get("import")).border = _thin_border()
        # Modificadores como texto concatenado
        mods = item.get("modifiers", [])
        ws_items.cell(row=i, column=8, value="\n".join(mods) if mods else "").border = _thin_border()
        ws_items.cell(row=i, column=8).alignment = Alignment(wrap_text=True, vertical="top")

    ws_items.column_dimensions["A"].width = 12
    ws_items.column_dimensions["B"].width = 25
    ws_items.column_dimensions["C"].width = 14
    ws_items.column_dimensions["D"].width = 8
    ws_items.column_dimensions["E"].width = 14
    ws_items.column_dimensions["F"].width = 16
    ws_items.column_dimensions["G"].width = 18
    ws_items.column_dimensions["H"].width = 40

    # ═══════════════════════════════════════════════════════════════════
    #  HOJA 3: RESUMEN (1 fila por factura, para consolidar)
    # ═══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_properties.tabColor = "F59E0B"

    # Labels legibles para los campos (orden prioritario)
    _LABELS = {
        "invoice_number": "N° Factura", "invoice_letter": "Tipo",
        "point_of_sale": "Punto de Venta",
        "emitter_name": "Emisor", "emitter_cuit": "CUIT Emisor",
        "emitter_address": "Dom. Emisor", "emitter_iibb": "IIBB Emisor",
        "emitter_iva_condition": "IVA Emisor",
        "receptor_name": "Receptor", "receptor_cuit": "CUIT Receptor",
        "receptor_address": "Dom. Receptor", "receptor_iibb": "IIBB Receptor",
        "receptor_account": "Cuenta", "receptor_deudor_account": "Cta. Deudora",
        "business_name": "Razón Social", "cuit": "CUIT",
        "iva_condition": "Condición IVA",
        "emission_date": "Fecha Emisión", "cae": "CAE",
        "invoice_cae": "CAE (pie)", "cae_expiry": "Vto. CAE",
        "invoice_type": "Tipo Factura",
        "incoterms": "Incoterms", "sap_number": "N° SAP",
        "oc_number": "OC", "payment_terms": "Condiciones de Pago",
        "shipping_method": "Forma de Envío",
        "road_company": "Empresa Vial", "road_cuit": "CUIT Vial",
        "period_start": "Período Desde", "period_end": "Período Hasta",
        "first_due_date": "1er Vto.", "first_due_amount": "1er Vto. $",
        "second_due_date": "2do Vto.", "second_due_amount": "2do Vto. $",
        "payment_method": "Medio de Pago", "client_code": "Cod. Cliente",
        "subtotal": "Subtotal", "importe_neto": "Importe Neto",
        "financiacion": "Financiación", "icl_amount": "ICL",
        "idc_amount": "IDC",
        "iva_inscripto": "IVA Inscripto", "iva_no_inscripto": "IVA No Insc.",
        "iva_percepcion": "IVA Percepción", "iva_percentage": "% IVA",
        "ingresos_brutos": "Ing. Brutos", "tasa_vial": "Tasa Vial",
        "net": "Neto Gravado", "iva_amount": "IVA",
        "total": "TOTAL",
    }

    # Agregar todos los campos que tengan valor (ordenados por _LABELS primero, luego alfabético)
    present_keys = [k for k in fields if fields[k].get("value") is not None and fields[k]["value"] != ""]
    ordered = [k for k in _LABELS if k in present_keys]
    ordered += sorted(k for k in present_keys if k not in ordered)
    headers = [_LABELS.get(k, k) for k in ordered]
    _set_header(ws2, 1, headers)
    for col, key in enumerate(ordered, 1):
        val = fields[key].get("value")
        ws2.cell(row=2, column=col, value=val).border = _thin_border()
    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 20


# ── Tablas ───────────────────────────────────────────────────────────────

def _write_tables(wb: Workbook, ext: Extraction) -> None:
    tables = ext.data.get("tables", [])
    if not tables:
        _write_generic(wb, None, ext)
        return

    for i, tbl in enumerate(tables):
        ws = wb.active if i == 0 else wb.create_sheet(f"Tabla {i + 1}")
        if i == 0:
            ws.title = "Tabla 1"
        rows = tbl.get("rows", [])
        if not rows:
            continue
        _set_header(ws, 1, rows[0])
        for r_idx, row in enumerate(rows[1:], 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val).border = _thin_border()
        for c in range(1, len(rows[0]) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18


# ── Contrato ─────────────────────────────────────────────────────────────

def _write_contract(wb: Workbook, ext: Extraction) -> None:
    ws = wb.active
    ws.title = "Texto"
    fields = ext.data.get("fields", {})
    meta = ext.data.get("meta", {})
    title = fields.get("title", {}).get("value", "Documento")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    full_text = meta.get("full_text", "")
    for i, line in enumerate(full_text.split("\n"), 3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


# ── Genérico ─────────────────────────────────────────────────────────────

def _write_generic(wb: Workbook, doc: Document | None, ext: Extraction | None) -> None:
    ws = wb.active
    ws.title = "Datos"
    if ext:
        fields = ext.data.get("fields", {})
        ws.cell(row=1, column=1, value="Campo").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Valor").font = Font(bold=True)
        row = 2
        for key, field in fields.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=field.get("value"))
            row += 1
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
